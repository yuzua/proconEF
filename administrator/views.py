from django.shortcuts import render
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.views.generic import TemplateView
from django .shortcuts import redirect
from accounts .models import CustomUser
from parking_req .models import ParkingUserModel
from carsharing_req .models import CarsharUserModel
from .forms import AdminParkingForm, UploadFileForm
from .models import MediaModel, StationModel, StationParkingModel
from owners_req .models import HostUserModel, CarInfoModel, ParentCategory, Category, CarInfoParkingModel
from owners_req .forms import CarInfoForm, CarOptionForm, CarInfoParkingForm, CarsharingDateForm
from survey .models import AnswerModel
from django.contrib import messages
from django.views import generic
import datetime
import json
import openpyxl
import re
import os, shutil
import ast
from django.contrib.auth.hashers import make_password


# Create your views here.

def index(request):
    params = {
        'hoge': '',
    }
    return render(request, 'administrator/index.html', params)


def test_ajax_response(request):
    input_lat = request.POST.getlist("name_input_lat")
    input_lng = request.POST.getlist("name_input_lng")
    hoge = "lat: "  + input_lat[0] + "lng: " + input_lng[0] + "がセットされました"
    request.session['user_lat'] = input_lat[0]
    request.session['user_lng'] = input_lng[0]

    return HttpResponse(hoge)

class ParkingAdminCreate(TemplateView):
    def __init__(self):
        self.params = {
            'title': 'ステーション追加',
            'message': '駐車場情報入力',
            'form': AdminParkingForm(),
        }
    
    def get(self, request):
        if str(request.user) == "AnonymousUser":
            print('ゲスト')
            messages.error(self.request, 'ログインしてください。')
            return redirect(to='/carsharing_req/index')
        else:
            self.params['lat'] = request.session['user_lat']
            self.params['lng'] = request.session['user_lng']
            return render(request, 'administrator/create.html', self.params)

    def post(self, request):
        obj = ParkingUserModel()
        parking = AdminParkingForm(request.POST, instance=obj)
        self.params['form'] = parking
        #バリデーションチェック
        if (parking.is_valid()):
            address = request.POST['address']
            parking_type = request.POST['parking_type']
            ground_type = request.POST['ground_type']
            width = request.POST['width']
            length = request.POST['length']
            height = request.POST['height']
            count = request.POST['count']
            # 確認画面へ
            data = {
                "address": address,
                "parking_type": parking_type,
                "ground_type": ground_type,
                "width": width,
                "length": length,
                "height": height,
                "count": count
            }
            self.params['data'] = data
            return render(request, 'administrator/checkparking.html', self.params)
        else:
            self.params['form'] = form
            messages.error(self.request, '入力データに問題があります')
        return render(request, 'administrator/create.html', self.params)

def checkparking(request):
    if (request.method == 'POST'):
        dt_now = datetime.datetime.now()
        user_id = 0
        address = request.POST['address']
        lat = request.session['user_lat']
        lng = request.session['user_lng']
        day = dt_now
        parking_type = request.POST['parking_type']
        ground_type = request.POST['ground_type']
        width = request.POST['width']
        length = request.POST['length']
        height = request.POST['height']
        count = request.POST['count']
        record = ParkingUserModel(user_id=user_id, address=address, lat=lat, lng=lng, day=day, \
            parking_type=parking_type, ground_type=ground_type, width=width, length=length, height=height, count=count)
        record.save()
        parking_obj = ParkingUserModel.objects.get(user_id=user_id, address=address, lat=lat, lng=lng, day=day, \
            parking_type=parking_type, ground_type=ground_type, width=width, length=length, height=height)
        ParkingSetStationArea(parking_obj.id, lat, lng)
        
        #セッションデータ削除
        del request.session['user_lat']
        del request.session['user_lng']
        if 'info_flag' in request.session:
            print(request.session['info_flag'])
            messages.success(request, '駐車場登録が完了しました。引き続き貸し出し車両・駐車場を選択してください。')
            return redirect(to='/administrator/settinginfo')
        else:
            print('none')
            messages.success(request, '駐車場登録が完了しました。')
            return redirect(to='/administrator/')
    else:
        messages.error(request, '不正なリクエストです。')
    return redirect(to='administrator:admin_main')

def ParkingSetStationArea(parking_id, lat, lng):
    item_list = list(StationModel.objects.values("id", "address", "lat", "lng"))
    numdict = {}
    for index in item_list:
        w = float(index['lat']) - float(lat)
        h = float(index['lng']) - float(lng)
        if w < 0:
            w = -w
        if h < 0:
            h = -h
        num = w + h
        numdict[index['id']] = num
    min_k = min(numdict, key=numdict.get)
    record = StationParkingModel(parking_id_id=parking_id, station_id_id=min_k)
    record.save()

def admin_main(request):
    s_p = ParkingUserModel.objects.filter(user_id=0)
    admin_parking = s_p.values("id", "address", "user_id", "lat", "lng")
    if (request.method == 'POST'):
        #num = request.POST['obj.id']
        num1 = request.POST['command']
        #edit
        if (num1 == 'edit'):
            num = request.POST['obj.id']
            obj = ParkingUserModel.objects.get(id=num)
            params = {
            'title': 'ParkingAdminEdit',
            'message': '駐車場情報修正',
            'id':num,
            'form': AdminParkingForm(instance=obj), 
            }
            return render(request, 'administrator/edit.html', params)
        #delete    
        if (num1 == 'delete'):
            num = request.POST['obj.id']
            delete_parking = ParkingUserModel.objects.get(id=num)
            params = {
            'title': 'ParkingAdminDelete',
            'message': '駐車場情報削除',
            'obj': delete_parking,
            'id': num,
            }
            return render(request, 'administrator/delete.html', params)
        #map
        if (num1 == 'map'):
            # data = CarsharUserModel.objects.get(id=request.session['user_id'])
            # print(data.pref01+data.addr01+data.addr02)
            add = '千葉県柏市末広町10-1'
            markerData = list(admin_parking.all())
            data = {
                'markerData': markerData,
            }
            params = {
            'name': '自宅',
            'add': add,
            'data_json': json.dumps(data)
            }
            if (request.method == 'POST'):
                params['add'] = request.POST['add']
            return render(request, "administrator/admin_main.html", params)    
    else:
        # data = CarsharUserModel.objects.get(id=0)
        # print(data.pref01+data.addr01+data.addr02)
        add = '千葉県柏市末広町10-1'
        #admin_main
        markerData = list(admin_parking.all())
        data = {
            'markerData': markerData,
        }
        params = {
            'title': 'ParkingAdminMain',
            'message': '駐車場登録データ',
            'data': admin_parking,
            'name': '自宅',
            'add': add,
            'data_json': json.dumps(data)
        }
        return render(request, 'administrator/admin_main.html', params)        

def edit(request):
    if (request.method == 'POST'):
        num = request.POST['p_id']
        obj = ParkingUserModel.objects.get(id=num)
        parking = AdminParkingForm(request.POST, instance=obj)
        params = {
            'title':'ParkingAdminEdit',
            'message': '駐車場情報修正', 
            'form': AdminParkingForm(),
            'id':num,
        }
        if (parking.is_valid()):
            parking.save()
            messages.success(request, '駐車場情報を修正しました。')
            return redirect(to='/administrator/')
        else:
            params['form'] = AdminParkingForm(request.POST, instance= obj)        
        
    return render(request, 'administrator/edit.html', params)

def delete(request, num):
    parking = ParkingUserModel.objects.get(id=num)
    if (request.method == 'POST'):
        parking.delete()
        messages.success(request, '駐車場情報を削除しました。')
        return redirect(to='/administrator/')

    return render(request, 'administrator/delete.html')

class CreateCarAdminView(TemplateView):
    def __init__(self):
        self.params = {
            'parentcategory_list': list(ParentCategory.objects.all()),
            'category_set': list(Category.objects.all().values()),
            'title': '車情報登録',
            'message': '車情報入力',
            'form': CarInfoForm(),
            'form2': CarOptionForm()
        }

    def post(self, request):
        obj = CarInfoModel()
        form = CarInfoForm(request.POST, instance=obj)
        form2 = CarOptionForm(request.POST, instance=obj)
        self.params['form'] = form
        self.params['form2'] = form2
        if form.is_valid() and form2.is_valid():
            car_maker = list(ParentCategory.objects.filter(id=request.POST['parent_category']).values("parent_category"))
            car_model = list(Category.objects.filter(id=request.POST['category']).values("category"))

            # 確認画面へ
            data = {
                "parent_category": request.POST['parent_category'],
                "category": request.POST['category'],
                "license_plate_place": request.POST['license_plate_place'],
                "license_plate_type": request.POST['license_plate_type'],
                "license_plate_how": request.POST['license_plate_how'],
                "license_plate_num": request.POST['license_plate_num'],
                "model_id": request.POST['model_id'],
                "people": request.POST['people'],
                "tire": request.POST['tire'],
                "at_mt": request.POST['at_mt'],
                "babysheet": 'babysheet' in request.POST,
                "car_nav": 'car_nav' in request.POST,
                "etc": 'etc' in request.POST,
                "car_autonomous": 'car_autonomous' in request.POST,
                "around_view_monitor": 'around_view_monitor' in request.POST,
                "non_smoking": 'non_smoking' in request.POST,
                "used_mileage": request.POST['used_mileage'],
                "used_years": request.POST['used_years'],
                "vehicle_inspection_day": request.POST['vehicle_inspection_day']
            }
            self.params['data'] = data
            self.params['car_maker'] = car_maker[0]['parent_category']
            self.params['car_model'] = car_model[0]['category']
            return render(request, 'administrator/checkcar.html', self.params)
        else:
            messages.error(self.request, '入力データに問題があります')
        return render(request, 'administrator/create.html', self.params)


        
    def get(self, request):
        if str(request.user) == "AnonymousUser":
            print('ゲスト')
            messages.error(self.request, 'ログインしてください。')
            return redirect(to='/carsharing_req/index')
        else:
            print(self.params['parentcategory_list'])
            print(self.params['category_set'])
            print(request.user)
        return render(request, 'administrator/createCar.html', self.params)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['parentcategory_list'] = ParentCategory.objects.all()
        return context

def checkcar(request):
    params = {
        'title': '車情報登録確認画面'
    }
    if (request.method == 'POST'):
        dt_now = datetime.datetime.now()
        day = dt_now
        user_id = 0
        parent_category = ParentCategory.objects.get(id=request.POST['parent_category'])
        category = Category.objects.get(id=request.POST['category'])
        # Str型からBool型に変換
        babysheet = judgmentTF(request.POST['babysheet'])
        car_nav = judgmentTF(request.POST['car_nav'])
        etc = judgmentTF(request.POST['etc'])
        car_autonomous = judgmentTF(request.POST['car_autonomous'])
        around_view_monitor = judgmentTF(request.POST['around_view_monitor'])
        non_smoking = judgmentTF(request.POST['non_smoking'])
        
        record = CarInfoModel(user_id=user_id, parent_category=parent_category, category=category, \
            license_plate_place=request.POST['license_plate_place'], license_plate_type=request.POST['license_plate_type'], \
            license_plate_how=request.POST['license_plate_how'], license_plate_num=request.POST['license_plate_num'], \
            model_id=request.POST['model_id'], people=request.POST['people'], tire=request.POST['tire'], at_mt=request.POST['at_mt'], \
            babysheet=babysheet, car_nav=car_nav, etc=etc, car_autonomous=car_autonomous, around_view_monitor=around_view_monitor, non_smoking=non_smoking, \
            used_mileage=request.POST['used_mileage'], used_years=request.POST['used_years'], \
            vehicle_inspection_day=request.POST['vehicle_inspection_day'], img=request.FILES['img'], day=day, key_flag=True)
        record.save()
        messages.success(request, '車両の登録が完了しました。引き続き駐車場情報を追加してください。')
        request.session['info_flag'] = True
        # return redirect(to='administrator:admin_main')
    else:
        messages.error(request, '不正なリクエストです。')
    return redirect(to='administrator:admin_main')

# ----------------------------------------------------------------------------------------------------------
def judgmentTF(string):
    if string == "True":
        boolean = True
    else:
        boolean = False

    return boolean
# ----------------------------------------------------------------------------------------------------------


class SettingAdminInfo(TemplateView):
    def __init__(self):
        self.params = {
            'title':'駐車場、車両情報登録データ',
            'message': '駐車場、車両情報登録データ',
            'car_data': '',
            'parking_data': '',
            # 'data_json': '',
        }
    def get(self, request):
        exclude_car= []
        exclude_parking = []
        if str(request.user) == "AnonymousUser":
            print('ゲスト')
            messages.error(self.request, 'ログインしてください。')
            return redirect(to='/carsharing_req/index')
        else:
            set_list = CarInfoParkingModel.objects.filter(user_id=0).values("car_id", "parking_id")
            for obj in set_list.values("car_id"):
                for index in obj.values():
                    exclude_car.append(index)
            for obj in set_list.values("parking_id"):
                for index in obj.values():
                    exclude_parking.append(index)
            car_list = CarInfoModel.objects.filter(user_id=0).exclude(id__in=exclude_car)
            # parking_list = ParkingUserModel.objects.filter(user_id=0).exclude(id__in=exclude_parking)
            parking_list = ParkingUserModel.objects.filter(user_id=0, countflag=True)
            self.params['car_data'] = car_list
            self.params['parking_data'] = parking_list
            # item = parking_list.values("id", "user_id", "lat", "lng")
            # item_list = list(item.all())
            # data = {
            #     'markerData': item_list,
            # }
            # self.params['data_json'] = json.dumps(data)
        return render(request, 'administrator/settinginfo.html', self.params)

    def post(self, request):
        user_id = int(request.POST['user_id'])
        car_id = CarInfoModel.objects.get(id=request.POST['car_id'])
        parking_id = ParkingUserModel.objects.get(id=request.POST['parking_id'])
        record = CarInfoParkingModel(user_id=user_id, car_id=car_id, parking_id=parking_id)
        record.save()
        print(request.POST['parking_id'])
        nowint = CarInfoParkingModel.objects.filter(parking_id=request.POST['parking_id']).count()
        maxint = ParkingUserModel.objects.filter(id=request.POST['parking_id']).values_list("count", flat=True)
        if nowint == maxint[0]:
            print('out')
            countflag = False
            ParkingUserModel.objects.filter(id=request.POST['parking_id']).update(countflag = countflag)
        else:
            print('safe')
        print(nowint)
        messages.success(self.request, '登録完了しました')
        return redirect(to='/administrator/')

def CreateSetting(request, num):
    params = {
        'title':'駐車場、車両情報登録データ(選択エリア)',
        'message': '駐車場、車両情報登録データ',
        'car_data': '',
        'parking_data': '',
        'area': num
    }
    if (request.method == 'POST'):
        user_id = int(request.POST['user_id'])
        car_id = CarInfoModel.objects.get(id=request.POST['car_id'])
        parking_id = ParkingUserModel.objects.get(id=request.POST['parking_id'])
        record = CarInfoParkingModel(user_id=user_id, car_id=car_id, parking_id=parking_id)
        record.save()
        print(request.POST['parking_id'])
        nowint = CarInfoParkingModel.objects.filter(parking_id=request.POST['parking_id']).count()
        maxint = ParkingUserModel.objects.filter(id=request.POST['parking_id']).values_list("count", flat=True)
        if nowint == maxint[0]:
            print('out')
            countflag = False
            ParkingUserModel.objects.filter(id=request.POST['parking_id']).update(countflag = countflag)
        else:
            print('safe')
        print(nowint)
        messages.success(request, '登録完了しました')
        return redirect(to='/administrator/')
    else:
        station_list = list(StationParkingModel.objects.filter(station_id_id=num).values("parking_id_id"))
        all_parking_list = []
        for parking_id in station_list:
            all_parking_list.append(parking_id['parking_id_id'])

        exclude_car= []
        exclude_parking = []
        if str(request.user) == "AnonymousUser":
            print('ゲスト')
            messages.error(request, 'ログインしてください。')
            return redirect(to='/carsharing_req/index')
        else:
            set_list = CarInfoParkingModel.objects.filter(user_id=0).values("car_id", "parking_id")
            for obj in set_list.values("car_id"):
                for index in obj.values():
                    exclude_car.append(index)
            for obj in set_list.values("parking_id"):
                for index in obj.values():
                    exclude_parking.append(index)
            car_list = CarInfoModel.objects.filter(user_id=0).exclude(id__in=exclude_car)
            parking_list = ParkingUserModel.objects.filter(user_id=0, countflag=True).filter(id__in=all_parking_list)
            print(parking_list)
            params['car_data'] = car_list
            params['parking_data'] = parking_list
        return render(request, 'administrator/settinginfo.html', params)


def DeleteSetting(request, num):
    if (request.method == 'POST'):
        print(request.POST['settinginfo'])
        record = CarInfoParkingModel.objects.get(id=request.POST['settinginfo'])
        record.delete()
        messages.success(request, '設置を解除しました。')
        return redirect(to='administrator:stationarea')
    else:
        station_list = list(StationParkingModel.objects.filter(station_id_id=num).values("parking_id_id"))
        all_parking_list = []
        for parking_id in station_list:
            all_parking_list.append(parking_id['parking_id_id'])
        # print(all_parking_list)
        item_list = list(CarInfoParkingModel.objects.filter(parking_id__in=all_parking_list).values())
        parking_list = []
        car_list = []
        for index in item_list:
            car_data = CarInfoModel.objects.get(id=index['car_id_id'])
            index['parent_category__parent_category'] = car_data.parent_category
            index['category__category'] = car_data.category
            index['img'] = car_data.img
            index['model_id'] = car_data.model_id
            parking_data = ParkingUserModel.objects.get(id=index['parking_id_id'])
            index['address'] = parking_data.address
            index['parking_type'] = parking_data.parking_type
            print(index)
            # parking_list.append(index['parking_id_id'])
            # car_list.append(index['car_id_id'])
        # parking_obj = list(ParkingUserModel.objects.filter(id__in=parking_list).values())
        # print(parking_obj)
        # print(item_list)
        # car_obj = list(CarInfoModel.objects.filter(id__in=car_list).select_related('parent_category__parent_category').select_related('category__category').values('id', 'user_id', 'parent_category__parent_category', 'category__category', 'model_id', 'people', 'tire', 'at_mt', 'babysheet', 'car_nav', 'etc', 'around_view_monitor', 'car_autonomous', 'non_smoking', 'used_mileage', 'used_years', 'img'))
        # for index, item in enumerate(item_list):
        #     for parking_item in parking_obj:
        #         if item['parking_id_id'] == parking_item['id']:
        #             parking_item['parking_id'] = parking_item['id']
        #             parking_item['id'] = item['id']
        # for index, item in enumerate(item_list):
        #     for car_item in car_obj:
        #         if item['car_id_id'] == car_item['id']:
        #             car_item['car_id'] = car_item['id']
        #             car_item['id'] = item['id']
            # parking_obj[index]['parking_id'] = parking_obj[index]['id']
            # parking_obj[index]['id'] = item['id']
            # car_obj[index]['car_id'] = car_obj[index]['id']
            # car_obj[index]['id'] = item['id']

    params = {
        'title': '配車設定解除',
        'data': item_list,
        'area': int(num)
    }
    return render(request, "administrator/deletesetting.html", params)


def superuser(request):
    superuser = CustomUser.objects.values()
    print(superuser)
    record = CustomUser()
    record.id = 0
    record.username = "admin"
    record.is_superuser = True
    record.is_staff = True
    record.is_active = True
    record.password = make_password("adminpass","WD2kFWBW5ya6")
    record.save()
    return redirect(to='/admin')

# データのダウンロード
class DownloadData(TemplateView):
    def __init__(self):
        self.params = {
            'title':'DownloadData',
            'path_list': ''
        }
    def get(self, request):
        DeleteUploadXlsx('./data/car_data/')
        DeleteUploadXlsx('./data/parking_data/')
        path_list = AllCarDownload()
        for index, item in enumerate(path_list):
            path_list[index] = item[7:]
        self.params['path_list'] = path_list
        return render(request, 'administrator/download_data.html', self.params)

# データのアップロード
class UploadData(TemplateView):
    def __init__(self):
        self.params = {
            'title':'UploadData',
            'form': ''
        }
    def get(self, request):
        form = UploadFileForm()
        self.params['form'] = form
        return render(request, 'administrator/upload_data.html', self.params)
    def post(self, request):
        file_name = str(request.FILES['attach'])
        result_json = re.match(r'.*\.json$', file_name)
        result_xlsx = re.match(r'.*\.xlsx$', file_name)
        # .jsonの場合
        if result_json:
            json_data = json.loads(request.POST['output'])
            if file_name == 'parentcategory.json':
                AllParentCategoryUpload(json_data)
                messages.success(self.request, 'メーカー情報をDBへ格納しました。')
            elif file_name == 'category.json':
                AllCategoryUpload(json_data)
                messages.success(self.request, '車種情報をDBへ格納しました。')
            elif file_name == 'settingcarparking.json':
                AllSettingCarUpload(json_data)
                messages.success(self.request, '車・駐車場設定をDBへ格納しました。')
            elif file_name == 'stationarea.json':
                AllStationAreaUpload(json_data)
                messages.success(self.request, 'ステーションエリア情報をDBへ格納しました。')
        # .xlsxの場合
        elif result_xlsx:
            #ファイルの場合はPOSTとFILEの両方を渡す
            #instanseを設定することでリクエストではなく、サーバー側で自動設定できる値を決められる
            mediaObj = MediaModel()
            administrator:UploadFileForm = UploadFileForm(request.POST, request.FILES, instance=mediaObj)
            administrator.errors
            # 値が正しいか確認(バリデーション)
            if administrator.is_valid():
                # 値をDBへの登録
                administrator.save()
            xlsx_all_list = ImportXlsx(file_name)
            if file_name[0] == 'p':
                AllParkingUpload(xlsx_all_list)
            else:
                AllCarUpload(xlsx_all_list)
            DeleteUploadXlsx('./media/xlsx/')
            messages.success(self.request, '車両情報をDBへ格納しました。')
        else:
            messages.error(self.request, 'error')
        return redirect(to='administrator:upload_data')




# -------------------------- DB上に追加保存された全車両のデータをxlsxに書き出し -------------------------- 
def AllCarDownload():
    dt_now = str(datetime.datetime.now())
    path_list = []

    # メーカー情報(親カテゴリー)をjsonでダウンロード
    pc_path = AllParentCategoryDownload()
    path_list.append(pc_path)
    # 車種情報(子カテゴリー)をjsonでダウンロード
    c_path = AllCategoryDownload()
    path_list.append(c_path)
    # 駐車場情報をExcelファイルで作成、ダウンロード
    parking_path = AllParkingDownload()
    path_list.append(parking_path)
    # 駐車場・車設定をExcelファイルで作成、ダウンロード
    settings_path = AllSettingDownload()
    path_list.append(settings_path)
    # 駐車場・ステーションエリア設定をExcelファイルで作成、ダウンロード
    station_path = AllStationAreaDownload()
    path_list.append(station_path)

    data = [
        ["車両ID","ユーザID","登録日","メーカー","車種","ナンバープレート-運輸支局-","ナンバープレート-車両種類-","ナンバープレート-使用用途-","ナンバープレート-指定番号-","型番","乗車人数","タイヤ","AT-MT","チャイルドシート","カーナビ","ETC","アラウンドビューモニター","自動運転","禁煙車","走行距離(km)","使用年数(年)","車検予定日","img","鍵工事"]
    ]
    data_list = list(CarInfoModel.objects.values())
    for data_dict in data_list:
        tmp = list(data_dict.values())
        # datetime型は入力できない為、str型にlist内を全変換：map()
        result = map((lambda x: str(x)), tmp)
        data.append(list(result))

    # ブックの読み込みとシート選択
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]

    # 行ごとに取り出してからExcelへ挿入
    for row in data:
        sheet.append(row)

    # xlsx型式で保存
    file_name = "./data/car_data/carinfo.xlsx"
    wb.save(file_name)
    path_list.append(file_name)
    return path_list

# ------------------------ DB上に追加保存された全駐車場のデータをxlsxに書き出し ------------------------ 
def AllParkingDownload():

    data = [
        ["駐車場ID","ユーザID","住所","緯度","経度","登録日","駐車場タイプ","土地タイプ","横幅","奥行き","高さ","収容台数","管理者","制限台数フラグ"]
    ]
    data_list = list(ParkingUserModel.objects.values())
    for data_dict in data_list:
        tmp = list(data_dict.values())
        # datetime型は入力できない為、str型にlist内を全変換：map()
        result = map((lambda x: str(x)), tmp)
        data.append(list(result))

    # ブックの読み込みとシート選択
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]

    # 行ごとに取り出してからExcelへ挿入
    for row in data:
        sheet.append(row)

    # xlsx型式で保存
    file_name = "./data/parking_data/parkinginfo.xlsx"
    wb.save(file_name)
    return file_name

# ---------------------- DB上に追加保存された全メーカー名(親カテゴリー)をjsonに書き出し ------------------------ 
def AllParentCategoryDownload():
    p_c = list(ParentCategory.objects.values())
    json_data = json.dumps(p_c, sort_keys=True, indent=4)

    # ファイルを開く(上書きモード)
    path = "./data/car_data/parentcategory.json"
    with open(path, 'w') as f:
        # jsonファイルの書き出し
        f.write(json_data)
    return path
# ------------------------- DB上に追加保存された全車種名(カテゴリー)をjsonに書き出し ------------------------------ 
def AllCategoryDownload():
    c = list(Category.objects.values())
    json_data = json.dumps(c, sort_keys=True, indent=4)

    # ファイルを開く(上書きモード)
    path = "./data/car_data/category.json"
    with open(path, 'w') as f:
        # jsonファイルの書き出し
        f.write(json_data)
    return path
# ---------------------------- DB上に追加保存された車・駐車場設定をjsonに書き出し -------------------------------- 
def AllSettingDownload():
    c = list(CarInfoParkingModel.objects.values())
    json_data = json.dumps(c, sort_keys=True, indent=4)

    # ファイルを開く(上書きモード)
    path = "./data/car_data/settingcarparking.json"
    with open(path, 'w') as f:
        # jsonファイルの書き出し
        f.write(json_data)
    return path
# --------------------- DB上に追加保存された駐車場・ステーションエリア設定をjsonに書き出し -------------------------- 
def AllStationAreaDownload():
    c = list(StationModel.objects.values())
    json_data = json.dumps(c, sort_keys=True, indent=4)

    # ファイルを開く(上書きモード)
    path = "./data/parking_data/stationarea.json"
    with open(path, 'w') as f:
        # jsonファイルの書き出し
        f.write(json_data)
    return path
# --------------------------------------------------------------------------------------------------------- 


# --------------------------------- 読み込んだjsonファイルをDBへ格納 ------------------------------------------- 
def AllParentCategoryUpload(json_data):
    print(type(json_data))
    flag = len(list(ParentCategory.objects.all()))
    if flag != 0:
        for data in json_data:
            record = ParentCategory.objects.get(id=data['id'])
            record.parent_category = data['parent_category']
            record.save()
    # 初回のみ
    else:
        for data in json_data:
            record = ParentCategory(id=data['id'], parent_category=data['parent_category'])
            record.save()

def AllCategoryUpload(json_data):
    print(json_data)
    flag = len(list(Category.objects.all()))
    if flag != 0:
        for data in json_data:
            record = Category.objects.get(id=data['id'])
            record.parent_category_id = data['parent_category_id']
            record.category = data['category']
            record.save()
    # 初回のみ
    else:
        for data in json_data:
            record = Category(id=data['id'], parent_category_id=data['parent_category_id'], category=data['category'])
            record.save()

def AllSettingCarUpload(json_data):
    print(json_data)
    flag = len(list(CarInfoParkingModel.objects.all()))
    if flag != 0:
        for data in json_data:
            record = CarInfoParkingModel.objects.get(id=data['id'])
            record.car_id_id = data['car_id_id']
            record.parking_id_id = data['parking_id_id']
            record.user_id = data['user_id']
            record.save()
    # 初回のみ
    else:
        record_list = []
        for data in json_data:
            record = CarInfoParkingModel(id=data['id'], car_id_id=data['car_id_id'], parking_id_id=data['parking_id_id'], user_id=data['user_id'])
            record_list.append(record)
            # record.save()
        CarInfoParkingModel.objects.bulk_create(record_list)

def AllStationAreaUpload(json_data):
    print(json_data)
    flag = len(list(StationModel.objects.all()))
    if flag != 0:
        for data in json_data:
            print(data)
            record = StationModel.objects.get(id=data['id'])
            record.address = data['address']
            record.lat = data['lat']
            record.lng = data['lng']
            record.save()
    # 初回のみ
    else:
        for data in json_data:
            record = StationModel(id=data['id'], address=data['address'], lat=data['lat'], lng=data['lng'])
            record.save()
        # 近いステーションを設定
        ParkingsSetStationArea()

# 駐車場の緯度経度から最寄りのステーションを設定(全件)
def ParkingsSetStationArea():
    result_list = {}
    item_list = list(StationModel.objects.values("id", "address", "lat", "lng"))
    item_list2 = list(ParkingUserModel.objects.values("id", "address", "lat", "lng"))
    for index2 in item_list2:
        lat = index2['lat']
        lng = index2['lng']
        numdict = {}
        for index in item_list:
            w = float(index['lat']) - float(lat)
            h = float(index['lng']) - float(lng)
            if w < 0:
                w = -w
            if h < 0:
                h = -h
            num = w + h
            numdict[index['id']] = num
        min_k = min(numdict, key=numdict.get)
        min_v = min(numdict.values())
        result_list[index2['id']] = min_k
    print(result_list)
    record_list = []
    for parking_id, station_id in result_list.items():
        print(parking_id)
        print(station_id)
        record = StationParkingModel(parking_id_id=parking_id, station_id_id=station_id)
        record_list.append(record)
    StationParkingModel.objects.bulk_create(record_list)
# --------------------------------- 読み込んだjsonファイルをDBへ格納 ---------------------------------- 
# ----------------------------- 読み込んだxlsxファイルの情報をを配列へ格納 ------------------------------
def ImportXlsx(file_name):
    # 今取り込んだxlsxファイルのpathを取得
    media = MediaModel.objects.values('attach')
    file_name = file_name.replace(' ', '_')
    file_name = file_name.replace(':', '')
    for path in media:
        if path['attach'][5:] == file_name:
            media_path = path['attach']
    
    # 取得したpathからファイルを開く
    wb = openpyxl.load_workbook("./media/" + media_path)
    # シートの名前を取得
    sheet_name = wb.sheetnames[0]
    sheet = wb[sheet_name]
    # シートの行数を取得
    m_row = sheet.max_row
    # シートの項目数を取得
    m_column = sheet.max_column
    # 全件取り出し
    l_2d = get_list_2d(sheet, 1, m_row, 1, m_column)
    print(l_2d)
    return l_2d

def get_list_2d(sheet, start_row, end_row, start_col, end_col):
    return get_value_list(sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col))

def get_value_list(t_2d):
    return([[cell.value for cell in row] for row in t_2d])
# ----------------------------- 読み込んだxlsxファイルの情報をを配列へ格納 ------------------------------

# ---------------------------------------- 車情報をDBへ保存 -----------------------------------------
def AllCarUpload(all_list):
    print(all_list[0])
    # 先頭のindexを削除
    all_list.pop(0)
    flag = len(list(CarInfoModel.objects.all()))
    if flag != 0:
        record_list = []
        for car_list in all_list:
            print(int(car_list[0]))
            if int(car_list[0]) > flag:
                record = CarInfoModel()
                record.id = int(car_list[0])
                record.user_id = int(car_list[1])
                # record.day = datetime.datetime.strptime(car_list[2], '%Y-%m-%d')
                record.day = car_list[2]
                record.parent_category_id = int(car_list[3])
                record.category_id = int(car_list[4])
                record.license_plate_place = car_list[5]
                record.license_plate_type = car_list[6]
                record.license_plate_how = car_list[7]
                record.license_plate_num = car_list[8]
                record.model_id = car_list[9]
                record.people = int(car_list[10])
                record.tire = car_list[11]
                record.at_mt = car_list[12]
                record.babysheet = car_list[13]
                record.car_nav = car_list[14]
                record.etc = car_list[15]
                record.around_view_monitor = car_list[16]
                record.car_autonomous = car_list[17]
                record.non_smoking = car_list[18]
                record.used_mileage = car_list[19]
                record.used_years = int(car_list[20])
                # record.vehicle_inspection_day = datetime.datetime.strptime(car_list[21], '%Y-%m-%d')
                record.vehicle_inspection_day = car_list[21]
                record.img = car_list[22]
                record.key_flag = car_list[23]
                record_list.append(record)
                # record.save()
        CarInfoModel.objects.bulk_create(record_list)
        print(record_list)
            # else:
            #     record = CarInfoModel.objects.get(id=car_list[0])
            #     record.user_id = int(car_list[1])
            #     record.day = datetime.datetime.strptime(car_list[2], '%Y-%m-%d')
            #     record.parent_category_id = int(car_list[3])
            #     record.category_id = int(car_list[4])
            #     record.license_plate_place = car_list[5]
            #     record.license_plate_type = car_list[6]
            #     record.license_plate_how = car_list[7]
            #     record.license_plate_num = car_list[8]
            #     record.model_id = car_list[9]
            #     record.people = int(car_list[10])
            #     record.tire = car_list[11]
            #     record.at_mt = car_list[12]
            #     record.babysheet = car_list[13]
            #     record.car_nav = car_list[14]
            #     record.etc = car_list[15]
            #     record.around_view_monitor = car_list[16]
            #     record.car_autonomous = car_list[17]
            #     record.non_smoking = car_list[18]
            #     record.used_mileage = car_list[19]
            #     record.used_years = int(car_list[20])
            #     record.vehicle_inspection_day = datetime.datetime.strptime(car_list[21], '%Y-%m-%d')
            #     record.img = car_list[22]
            #     record.key_flag = car_list[23]
                # record.save()
    else:
        record_list = []
        for car_list in all_list:
            record = CarInfoModel()
            record.user_id = int(car_list[1])
            #record.day = datetime.datetime.strptime(car_list[2], '%Y-%m-%d')
            record.day = car_list[2]
            record.parent_category_id = int(car_list[3])
            record.category_id = int(car_list[4])
            record.license_plate_place = car_list[5]
            record.license_plate_type = car_list[6]
            record.license_plate_how = car_list[7]
            record.license_plate_num = car_list[8]
            record.model_id = car_list[9]
            record.people = int(car_list[10])
            record.tire = car_list[11]
            record.at_mt = car_list[12]
            record.babysheet = car_list[13]
            record.car_nav = car_list[14]
            record.etc = car_list[15]
            record.around_view_monitor = car_list[16]
            record.car_autonomous = car_list[17]
            record.non_smoking = car_list[18]
            record.used_mileage = car_list[19]
            record.used_years = int(car_list[20])
            #record.vehicle_inspection_day = datetime.datetime.strptime(car_list[21], '%Y-%m-%d')
            record.vehicle_inspection_day = car_list[21]
            record.img = car_list[22]
            record.key_flag = car_list[23]
            record_list.append(record)
            # record.save()
        CarInfoModel.objects.bulk_create(record_list)

# ---------------------------------------- 車情報をDBへ保存 -----------------------------------------

# --------------------------------------- 駐車場情報をDBへ保存 ---------------------------------------
def AllParkingUpload(all_list):
    print(all_list[0])
    # 先頭のindexを削除
    all_list.pop(0)
    flag = len(list(ParkingUserModel.objects.all()))
    if flag != 0:
        for parking_list in all_list:
            record = ParkingUserModel.objects.get(id=parking_list[0])
            record.user_id = int(parking_list[1])
            record.address = parking_list[2]
            record.lat = parking_list[3]
            record.lng = parking_list[4]
            record.day = datetime.datetime.strptime(parking_list[5], '%Y-%m-%d')
            record.parking_type = parking_list[6]
            record.ground_type = parking_list[7]
            record.width = int(parking_list[8])
            record.length = int(parking_list[9])
            record.height = int(parking_list[10])
            record.count = int(parking_list[11])
            record.admin = parking_list[12]
            record.countflag = parking_list[13]
            record.save()
    else:
        record_list = []
        for parking_list in all_list:
            record = ParkingUserModel()
            record.user_id = int(parking_list[1])
            record.address = parking_list[2]
            record.lat = parking_list[3]
            record.lng = parking_list[4]
            record.day = datetime.datetime.strptime(parking_list[5], '%Y-%m-%d')
            # record.day = parking_list[5]
            record.parking_type = parking_list[6]
            record.ground_type = parking_list[7]
            record.width = int(parking_list[8])
            record.length = int(parking_list[9])
            record.height = int(parking_list[10])
            record.count = int(parking_list[11])
            record.admin = parking_list[12]
            record.countflag = parking_list[13]
            record_list.append(record)
            # record.save()
        ParkingUserModel.objects.bulk_create(record_list)

# --------------------------------------- 駐車場情報をDBへ保存 ---------------------------------------

# 引数に指定されたフォルダー内のファイルを削除
def DeleteUploadXlsx(target_dir):
    shutil.rmtree(target_dir)
    os.mkdir(target_dir)


def StationArea(request):
    add = '〒277-0005 千葉県柏市柏１丁目１−１'
    item_list = list(StationModel.objects.values("id", "address", "lat", "lng"))
    item_list2 = list(ParkingUserModel.objects.values("id", "address", "lat", "lng"))

    path = "./data/haishayosoku/result.json"
    haishayosoku_dict = openJSON(path)
    # print(haishayosoku_dict['ID'])
    # print(haishayosoku_dict['kategori'])
    # print(haishayosoku_dict['kazu'])
    for items in item_list:
        index = int(items['id']) - 1
        items['color'] = haishayosoku_dict['kategori'].get(str(index))
        items['many'] = haishayosoku_dict['kazu'].get(str(index))

    data = {
        'markerData': item_list,
        'markerData2': item_list2
    }
    params = {
        'title': 'エリア',
        'name': '自宅',
        'add': add,
        'data_json': json.dumps(data),
        'latlng': 'undefined'
    }
    if (request.method == 'POST'):
        params['title'] = '検索地付近で検索'
        params['add'] = request.POST['add']
        params['name'] = '検索'
    return render(request, "administrator/stationarea.html", params)

def openJSON(path):
    with open(path, 'r') as f:
        # jsonファイルの読み込み
        json_data = f.read()
        json_object = json.loads(json_data)

    return json_object



def mobile(request):
    set_list = CarInfoParkingModel.objects.values("parking_id")
    item_all = ParkingUserModel.objects.filter(id__in=set_list)
    item = item_all.values("id", "address", "user_id", "lat", "lng")
    item_list = list(item.all())
    data = {
        'markerData': item_list,
    }
    return HttpResponse(json.dumps(data))



def survey(request):
    params = {
        'title': "アンケート"
    }
    data = list(AnswerModel.objects.filter(count=3).values('answer'))
    for itemstr in data:
        item = ast.literal_eval(itemstr['answer'])
        print(type(item))
        print(item)
    return render(request, 'administrator/survey.html', params)